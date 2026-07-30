"""AI-assisted, dynamic update of the JS framework's real test data workbook
(test-data/TestData.xlsx) — the "use AI to manage and update it dynamically"
part of Phase 3.

Takes a natural-language instruction, shows the LLM the *actual* existing
headers and a couple of sample rows for the target sheet (so it matches the
real schema instead of inventing columns), asks it to propose one new row as
JSON, then appends it via openpyxl. Cucumber's step definitions
(src/utils/testDataHelper.js) read this same file by scenarioId, so a new
row here is immediately usable as `"NEW_SCENARIO_ID"` in a `.feature` file.

Usage:
    python ai_update_test_data.py --sheet TravellerDetails \
        --instruction "add a negative case with an empty passport number for an international booking" \
        --scenario-id TRAVELLER_DETAILS_MISSING_PASSPORT

    python ai_update_test_data.py --sheet TravellerDetails --instruction "..." --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import load_workbook

from common.llm import OllamaClient

TEST_DATA_PATH = REPO_ROOT / "test-data" / "TestData.xlsx"

ROW_PROPOSAL_SYSTEM_PROMPT = """You are a QA engineer adding a new row to an existing test data \
spreadsheet. You are given the sheet's real column headers and a few real existing rows as \
examples, plus a natural-language instruction for the new row. Reply with ONLY a JSON object \
whose keys are exactly the given headers (same spelling/case, all of them) and whose values are \
the new row's data, consistent in type/format with the example rows. No prose, no markdown code \
fence, just the JSON object."""


def read_sheet_schema(sheet_name: str, sample_rows: int = 3):
    wb = load_workbook(TEST_DATA_PATH)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f'Sheet "{sheet_name}" not found. Available: {", ".join(wb.sheetnames)}')
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    samples = [dict(zip(headers, r)) for r in rows[1 : 1 + sample_rows]]
    return wb, ws, headers, samples


def propose_row(client: OllamaClient, headers: list[str], samples: list[dict], instruction: str) -> dict:
    user_msg = (
        f"HEADERS: {json.dumps(headers)}\n\n"
        f"EXAMPLE ROWS:\n{json.dumps(samples, indent=2, default=str)}\n\n"
        f"INSTRUCTION: {instruction}"
    )
    raw = client.ask(ROW_PROPOSAL_SYSTEM_PROMPT, user_msg)
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        raise RuntimeError(f"Model did not return a JSON object. Raw response:\n{raw}")
    row = json.loads(match.group(0))
    missing = [h for h in headers if h not in row]
    if missing:
        raise RuntimeError(f"Proposed row is missing header(s) {missing}. Raw: {row}")
    return row


def main():
    parser = argparse.ArgumentParser(description="AI-assisted dynamic test data updater")
    parser.add_argument("--sheet", required=True, help="Target sheet name in test-data/TestData.xlsx")
    parser.add_argument("--instruction", required=True, help="Natural-language description of the new row")
    parser.add_argument("--scenario-id", help="Override the generated scenarioId (if the sheet has one)")
    parser.add_argument("--dry-run", action="store_true", help="Preview the proposed row without writing it")
    args = parser.parse_args()

    print(f"Reading schema for sheet '{args.sheet}' from {TEST_DATA_PATH}...")
    wb, ws, headers, samples = read_sheet_schema(args.sheet)
    print(f"Headers: {headers}")

    client = OllamaClient()
    print("Asking local LLM to propose a new row...")
    row = propose_row(client, headers, samples, args.instruction)

    if args.scenario_id and "scenarioId" in row:
        row["scenarioId"] = args.scenario_id

    print("\nProposed row:")
    for h in headers:
        print(f"  {h}: {row[h]}")

    if args.dry_run:
        print("\n--dry-run set: not written.")
        return

    ws.append([row[h] for h in headers])
    wb.save(TEST_DATA_PATH)
    print(f"\nAppended to '{args.sheet}' in {TEST_DATA_PATH}.")


if __name__ == "__main__":
    main()
